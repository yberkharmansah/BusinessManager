# modules/supplier_tab.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import (
    get_all_suppliers, add_supplier, update_supplier, delete_supplier,
    get_supplier_balances, add_smart_balance_transaction, get_supplier_total_balance,
    get_due_supplier_balances, fetch_one, get_supplier_transaction_history
)

class SupplierTab:
    def __init__(self, parent, branch_id):
        self.parent = parent
        self.branch_id = branch_id
        
        # Sıralama için değişkenler
        self.sort_column = None
        self.sort_reverse = False
        self.selected_supplier_id = None
        
        self.create_widgets()
        self.load_suppliers()
        self.load_balances()
    
    def create_widgets(self):
        # Üst çerçeve - Toptancı Yönetimi
        supplier_frame = tk.LabelFrame(self.parent, text="🤝 Toptancı Yönetimi", padx=10, pady=10)
        supplier_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Toptancı Butonları
        tk.Button(
            supplier_frame,
            text="➕ Toptancı Ekle",
            command=self.add_supplier_dialog,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            supplier_frame,
            text="⚠️ Yaklaşan Ödemeler",
            command=self.show_due_payments,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            supplier_frame,
            text="📊 Excel Raporu",
            command=self.export_to_excel,
            bg="#2196F3",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.RIGHT, padx=5)
        
        # Toptancı Listesi
        list_frame = tk.Frame(self.parent)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Sol panel - Toptancılar
        left_panel = tk.Frame(list_frame)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        tk.Label(left_panel, text="📋 Toptancı Listesi", font=("Arial", 12, "bold")).pack(pady=5)
        
        # Scrollbar
        scroll_y = tk.Scrollbar(left_panel)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.supplier_tree = ttk.Treeview(
            left_panel,
            columns=("ID", "Name", "Type", "Phone", "TotalBalance"),
            show="headings",
            height=15,
            yscrollcommand=scroll_y.set
        )
        
        # Sıralama başlıkları
        self.supplier_tree.heading("ID", text="ID", command=lambda: self.sort_supplier("ID"))
        self.supplier_tree.heading("Name", text="Toptancı", command=lambda: self.sort_supplier("Name"))
        self.supplier_tree.heading("Type", text="Tür", command=lambda: self.sort_supplier("Type"))
        self.supplier_tree.heading("Phone", text="Telefon", command=lambda: self.sort_supplier("Phone"))
        self.supplier_tree.heading("TotalBalance", text="Toplam Bakiye", command=lambda: self.sort_supplier("TotalBalance"))
        
        self.supplier_tree.column("ID", width=50, anchor="center")
        self.supplier_tree.column("Name", width=200)
        self.supplier_tree.column("Type", width=100)
        self.supplier_tree.column("Phone", width=120)
        self.supplier_tree.column("TotalBalance", width=100, anchor="e")
        
        self.supplier_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y.config(command=self.supplier_tree.yview)
        
        # Çift tıklama ve sağ tık menüsü
        self.supplier_menu = tk.Menu(self.supplier_tree, tearoff=0)
        self.supplier_menu.add_command(label="✏️ Düzenle", command=self.edit_supplier)
        self.supplier_menu.add_command(label="🗑️ Sil", command=self.delete_supplier)
        self.supplier_menu.add_separator()
        self.supplier_menu.add_command(label="💵 Ödeme Yap", command=self.payment_from_context)
        self.supplier_menu.add_command(label="➕ Borç Ekle", command=self.add_debt_from_context)
        self.supplier_menu.add_command(label="➕ Alacak Ekle", command=self.add_credit_from_context)
        self.supplier_menu.add_separator()
        self.supplier_menu.add_command(label="📋 Tüm Hareketler", command=self.show_all_transactions)
        
        self.supplier_tree.bind("<Button-3>", self.show_supplier_menu)
        self.supplier_tree.bind("<Double-1>", lambda e: self.edit_supplier())
        
        # Sağ panel - Bakiyeler
        right_panel = tk.Frame(list_frame)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        tk.Label(right_panel, text="💰 Güncel Bakiyeler", font=("Arial", 12, "bold")).pack(pady=5)
        
        # Scrollbar
        scroll_y2 = tk.Scrollbar(right_panel)
        scroll_y2.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.balance_tree = ttk.Treeview(
            right_panel,
            columns=("BalanceID", "Date", "Supplier", "Type", "Amount", "DueDate", "Status", "Desc"),
            show="headings",
            height=15,
            yscrollcommand=scroll_y2.set
        )
        
        self.balance_tree.heading("BalanceID", text="ID")
        self.balance_tree.heading("Date", text="Tarih")
        self.balance_tree.heading("Supplier", text="Toptancı")
        self.balance_tree.heading("Type", text="İşlem")
        self.balance_tree.heading("Amount", text="Tutar")
        self.balance_tree.heading("DueDate", text="Vade")
        self.balance_tree.heading("Status", text="Durum")
        self.balance_tree.heading("Desc", text="Açıklama")
        
        self.balance_tree.column("BalanceID", width=50, anchor="center")
        self.balance_tree.column("Date", width=120)
        self.balance_tree.column("Supplier", width=140)
        self.balance_tree.column("Type", width=80, anchor="center")
        self.balance_tree.column("Amount", width=80, anchor="e")
        self.balance_tree.column("DueDate", width=100, anchor="center")
        self.balance_tree.column("Status", width=80, anchor="center")
        self.balance_tree.column("Desc", width=150)
        
        self.balance_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y2.config(command=self.balance_tree.yview)
        
        # Bakiye sağ tık menüsü
        self.balance_menu = tk.Menu(self.balance_tree, tearoff=0)
        self.balance_menu.add_command(label="💵 Ödeme Yap", command=self.payment_from_balance)
        self.balance_menu.add_command(label="🔄 Durum Güncelle", command=self.update_balance_status)
        self.balance_menu.add_separator()
        self.balance_menu.add_command(label="📋 Detaylar", command=self.show_balance_details)
        
        self.balance_tree.bind("<Button-3>", self.show_balance_menu)
        
        # Alt çerçeve - Bilgi
        bottom_frame = tk.Frame(self.parent)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.info_label = tk.Label(bottom_frame, text="Toptancı: 0 | Aktif Bakiye: 0 | Gecikmiş: 0", font=("Arial", 10))
        self.info_label.pack(side=tk.LEFT)
        
        tk.Button(
            bottom_frame,
            text="🔄 Yenile",
            command=lambda: [self.load_suppliers(), self.load_balances()],
            bg="#9E9E9E",
            fg="white"
        ).pack(side=tk.RIGHT)
    
    def sort_supplier(self, col):
        """Toptancı listesini sıralar"""
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False
        self.load_suppliers()
    
    def load_suppliers(self):
        """Toptancıları listeler"""
        for item in self.supplier_tree.get_children():
            self.supplier_tree.delete(item)
        
        suppliers = get_all_suppliers(self.branch_id)
        
        # Sıralama
        if self.sort_column:
            def sort_key(supplier):
                if self.sort_column == "ID":
                    return supplier['id']
                elif self.sort_column == "Name":
                    return supplier['name'].lower()
                elif self.sort_column == "Type":
                    return supplier['supplier_type'].lower()
                elif self.sort_column == "Phone":
                    return supplier['phone'] or ""
                elif self.sort_column == "TotalBalance":
                    return get_supplier_total_balance(supplier['id'])
                return 0
            
            suppliers.sort(key=sort_key, reverse=self.sort_reverse)
        
        # Toptancıları ekle
        for supplier in suppliers:
            total_balance = get_supplier_total_balance(supplier['id'])
            
            # Renklendirme
            if total_balance > 0:
                tag = 'positive'  # Alacağımız var
            elif total_balance < 0:
                tag = 'negative'  # Borcumuz var
            else:
                tag = 'neutral'
            
            self.supplier_tree.insert("", tk.END, values=(
                supplier['id'],
                supplier['name'],
                supplier['supplier_type'],
                supplier['phone'] or "-",
                f"₺{total_balance:.2f}"
            ), tags=(tag,))
        
        # Renk ayarları
        self.supplier_tree.tag_configure('positive', background='#e8f5e9', foreground='#2e7d32')
        self.supplier_tree.tag_configure('negative', background='#ffebee', foreground='#c62828')
        self.supplier_tree.tag_configure('neutral', background='#f5f5f5', foreground='#616161')
        
        self.update_info_label()
    
    def load_balances(self):
        """Bakiye hareketlerini listeler"""
        for item in self.balance_tree.get_children():
            self.balance_tree.delete(item)
        
        balances = get_supplier_balances(self.branch_id)
        
        for balance in balances:
            # Durum renklendirme
            status = balance['status']
            due_date = balance['due_date']
            balance_type = balance['balance_type']
            
            if status == 'GECIKMIS':
                tag = 'late'
            elif status == 'AKTIF' and due_date:
                days_left = (datetime.strptime(due_date, "%Y-%m-%d") - datetime.now()).days
                if days_left < 0:
                    tag = 'late'
                elif days_left <= 3:
                    tag = 'warning'
                elif days_left <= 7:
                    tag = 'caution'
                else:
                    tag = 'active'
            elif status == 'ODENDI':
                tag = 'paid'
            else:
                tag = 'active'
            
            self.balance_tree.insert("", tk.END, values=(
                balance['id'],
                balance['date'],
                balance['supplier_name'],
                balance_type,
                f"₺{balance['amount']:.2f}",
                due_date or "-",
                status,
                balance['description'] or "-"
            ), tags=(tag,))
        
        # Renk ayarları
        self.balance_tree.tag_configure('late', background='#ffebee', foreground='#c62828')
        self.balance_tree.tag_configure('warning', background='#fff3e0', foreground='#e65100')
        self.balance_tree.tag_configure('caution', background='#fff9c4', foreground='#f57f17')
        self.balance_tree.tag_configure('active', background='#e8f5e9', foreground='#2e7d32')
        self.balance_tree.tag_configure('paid', background='#f5f5f5', foreground='#616161')
    
    def update_info_label(self):
        """Bilgi etiketini günceller"""
        suppliers = get_all_suppliers(self.branch_id)
        balances = get_supplier_balances(self.branch_id)
        
        active_balances = len([b for b in balances if b['status'] == 'AKTIF'])
        overdue_balances = len([b for b in balances if b['status'] == 'GECIKMIS'])
        
        self.info_label.config(
            text=f"Toptancı: {len(suppliers)} | Aktif Bakiye: {active_balances} | Gecikmiş: {overdue_balances}"
        )
    
    def show_supplier_menu(self, event):
        """Toptancı sağ tık menüsü"""
        # Seçili öğeyi belirle
        item = self.supplier_tree.identify_row(event.y)
        if item:
            self.supplier_tree.selection_set(item)
            self.selected_supplier_id = self.supplier_tree.item(item)['values'][0]
            self.supplier_menu.post(event.x_root, event.y_root)
    
    def show_balance_menu(self, event):
        """Bakiye sağ tık menüsü"""
        item = self.balance_tree.identify_row(event.y)
        if item:
            self.balance_tree.selection_set(item)
            self.balance_menu.post(event.x_root, event.y_root)
    
    def add_supplier_dialog(self):
        """Toptancı ekleme penceresi - TÜR TEXTBOX"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("➕ Yeni Toptancı Ekle")
        dialog.geometry("400x450")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # Form
        tk.Label(dialog, text="Toptancı Adı *:").pack(pady=5)
        name_entry = tk.Entry(dialog, width=30, font=("Arial", 11))
        name_entry.pack()
        
        tk.Label(dialog, text="Türü * (Örn: Meşrubat, Süt, Yemek):").pack(pady=5)
        type_entry = tk.Entry(dialog, width=30, font=("Arial", 11))  # TEXTBOX OLDU!
        type_entry.insert(0, "Meşrubat")  # Varsayılan
        type_entry.pack()
        
        tk.Label(dialog, text="Telefon:").pack(pady=5)
        phone_entry = tk.Entry(dialog, width=30, font=("Arial", 11))
        phone_entry.pack()
        
        tk.Label(dialog, text="E-posta:").pack(pady=5)
        email_entry = tk.Entry(dialog, width=30, font=("Arial", 11))
        email_entry.pack()
        
        # Butonlar
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        tk.Button(
            button_frame,
            text="💾 Kaydet",
            command=lambda: self.save_supplier(dialog, name_entry, type_entry, phone_entry, email_entry),
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
        
        tk.Button(
            button_frame,
            text="❌ İptal",
            command=dialog.destroy,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
    
    def save_supplier(self, dialog, name_entry, type_entry, phone_entry, email_entry):
        """Toptancıyı kaydeder"""
        name = name_entry.get().strip()
        supplier_type = type_entry.get().strip()  # TEXTBOX DEĞERİ
        phone = phone_entry.get().strip()
        email = email_entry.get().strip()
        
        if not name or not supplier_type:
            messagebox.showwarning("Eksik Bilgi", "Ad ve tür zorunludur!")
            return
        
        supplier_id = add_supplier(self.branch_id, name, supplier_type, phone, email)
        
        if supplier_id:
            messagebox.showinfo("Başarılı", f"Toptancı eklendi: {name}")
            dialog.destroy()
            self.load_suppliers()
        else:
            messagebox.showerror("Hata", "Toptancı eklenemedi!")
    
    def edit_supplier(self):
        """Toptancı düzenleme - TÜR TEXTBOX"""
        selection = self.supplier_tree.selection()
        if not selection:
            messagebox.showwarning("Seçim Yok", "Lütfen bir toptancı seçin!")
            return
        
        supplier_id = self.supplier_tree.item(selection[0])['values'][0]
        supplier = fetch_one("SELECT * FROM suppliers WHERE id = ?", (supplier_id,))
        
        if not supplier:
            return
        
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"✏️ Toptancı Düzenle: {supplier['name']}")
        dialog.geometry("400x450")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # Form
        tk.Label(dialog, text="Toptancı Adı *:").pack(pady=5)
        name_entry = tk.Entry(dialog, width=30, font=("Arial", 11))
        name_entry.insert(0, supplier['name'])
        name_entry.pack()
        
        tk.Label(dialog, text="Türü * (Örn: Meşrubat, Süt, Yemek):").pack(pady=5)
        type_entry = tk.Entry(dialog, width=30, font=("Arial", 11))  # TEXTBOX OLDU!
        type_entry.insert(0, supplier['supplier_type'])
        type_entry.pack()
        
        tk.Label(dialog, text="Telefon:").pack(pady=5)
        phone_entry = tk.Entry(dialog, width=30, font=("Arial", 11))
        phone_entry.insert(0, supplier['phone'] or "")
        phone_entry.pack()
        
        tk.Label(dialog, text="E-posta:").pack(pady=5)
        email_entry = tk.Entry(dialog, width=30, font=("Arial", 11))
        email_entry.insert(0, supplier['email'] or "")
        email_entry.pack()
        
        # Butonlar
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        tk.Button(
            button_frame,
            text="💾 Güncelle",
            command=lambda: self.save_edited_supplier(dialog, supplier_id, name_entry, type_entry, phone_entry, email_entry),
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
        
        tk.Button(
            button_frame,
            text="❌ İptal",
            command=dialog.destroy,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
    
    def save_edited_supplier(self, dialog, supplier_id, name_entry, type_entry, phone_entry, email_entry):
        """Düzenlenmiş toptancıyı kaydeder"""
        name = name_entry.get().strip()
        supplier_type = type_entry.get().strip()  # TEXTBOX DEĞERİ
        phone = phone_entry.get().strip()
        email = email_entry.get().strip()
        
        if not name or not supplier_type:
            messagebox.showwarning("Eksik Bilgi", "Ad ve tür zorunludur!")
            return
        
        success = update_supplier(supplier_id, name, supplier_type, phone, email)
        
        if success:
            messagebox.showinfo("Başarılı", f"Toptancı güncellendi: {name}")
            dialog.destroy()
            self.load_suppliers()
        else:
            messagebox.showerror("Hata", "Toptancı güncellenemedi!")
    
    def delete_supplier(self):
        """Toptancıyı siler"""
        selection = self.supplier_tree.selection()
        if not selection:
            messagebox.showwarning("Seçim Yok", "Lütfen bir toptancı seçin!")
            return
        
        supplier_id = self.supplier_tree.item(selection[0])['values'][0]
        supplier_name = self.supplier_tree.item(selection[0])['values'][1]
        
        if messagebox.askyesno("Silme Onayı", f"'{supplier_name}' ve tüm bakiyeleri silinsin mi?"):
            success = delete_supplier(supplier_id)
            if success:
                messagebox.showinfo("Başarılı", "Toptancı silindi!")
                self.load_suppliers()
                self.load_balances()
            else:
                messagebox.showerror("Hata", "Toptancı silinemedi!")
    
    # SAĞ TIK İŞLEMLERİ
    def payment_from_context(self):
        """Sağ tık → Ödeme Yap"""
        if not self.selected_supplier_id:
            return
        
        supplier = fetch_one("SELECT * FROM suppliers WHERE id = ?", (self.selected_supplier_id,))
        current_balance = get_supplier_total_balance(self.selected_supplier_id)
        
        if current_balance >= 0:
            messagebox.showinfo("Bilgi", f"'{supplier['name']}' için ödeme yapacak borcunuz yok!\nMevcut bakiye: ₺{current_balance:.2f}")
            return
        
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"💵 Ödeme Yap: {supplier['name']}")
        dialog.geometry("400x350")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        tk.Label(dialog, text=f"Toptancı: {supplier['name']}", font=("Arial", 11, "bold")).pack(pady=10)
        tk.Label(dialog, text=f"Mevcut Borç: ₺{abs(current_balance):.2f}", font=("Arial", 10)).pack()
        
        tk.Label(dialog, text="Ödeme Tutarı:").pack(pady=10)
        amount_entry = tk.Entry(dialog, width=20, font=("Arial", 12))
        amount_entry.pack()
        
        tk.Label(dialog, text="Tarih:").pack(pady=5)
        date_entry = tk.Entry(dialog, width=20, font=("Arial", 11))
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        date_entry.pack()
        
        tk.Label(dialog, text="Açıklama:").pack(pady=5)
        desc_text = tk.Text(dialog, height=3, width=40)
        desc_text.insert(tk.END, "Ödeme yapıldı")
        desc_text.pack()
        
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        tk.Button(
            button_frame,
            text="💵 Ödeme Yap",
            command=lambda: self.process_smart_payment(dialog, self.selected_supplier_id, amount_entry, date_entry, desc_text, current_balance),
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
        
        tk.Button(
            button_frame,
            text="❌ İptal",
            command=dialog.destroy,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
    
    def process_smart_payment(self, dialog, supplier_id, amount_entry, date_entry, desc_text, current_balance):
        """Akıllı ödeme işlemini gerçekleştirir"""
        try:
            amount = float(amount_entry.get())
            payment_date = date_entry.get().strip()
            description = desc_text.get("1.0", tk.END).strip()
            
            # Tarih doğrulama
            try:
                datetime.strptime(payment_date, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning("Hatalı Tarih", "Tarih formatı: YYYY-MM-DD")
                return
            
            if amount <= 0:
                messagebox.showwarning("Hatalı Değer", "Ödeme miktarı 0'dan büyük olmalı!")
                return
            
            if amount > abs(current_balance):
                if not messagebox.askyesno("Fazla Ödeme", f"Ödeme miktarı borçtan fazla!\nFazla ödeme alacak olarak kaydedilsin mi?"):
                    return
            
            # Akıllı ödeme işlemi
            new_balance = add_smart_balance_transaction(
                supplier_id, 
                "ODEME", 
                amount, 
                due_date=None,
                description=description,
                transaction_date=payment_date
            )
            
            if new_balance is not None:
                supplier = fetch_one("SELECT * FROM suppliers WHERE id = ?", (supplier_id,))
                if new_balance > 0:
                    messagebox.showinfo("Başarılı", f"Ödeme tamamlandı!\nYeni alacağınız: ₺{new_balance:.2f}")
                elif new_balance < 0:
                    messagebox.showinfo("Başarılı", f"Ödeme tamamlandı!\nKalan borcunuz: ₺{abs(new_balance):.2f}")
                else:
                    messagebox.showinfo("Başarılı", "Ödeme tamamlandı!\nBakiye sıfırlandı.")
                
                dialog.destroy()
                self.load_balances()
                self.load_suppliers()
            else:
                messagebox.showerror("Hata", "Ödeme işlemi gerçekleştirilemedi!")
                
        except ValueError:
            messagebox.showwarning("Hatalı Değer", "Miktar sayısal olmalı!")
    
    def add_debt_from_context(self):
        """Sağ tık → Borç Ekle"""
        self.add_balance_dialog("BORC")
    
    def add_credit_from_context(self):
        """Sağ tık → Alacak Ekle"""
        self.add_balance_dialog("ALACAK")
    
    def add_balance_dialog(self, transaction_type="BORC"):
        """Bakiye ekleme penceresi - Tür seçmeli"""
        if not self.selected_supplier_id:
            return
        
        supplier = fetch_one("SELECT * FROM suppliers WHERE id = ?", (self.selected_supplier_id,))
        current_balance = get_supplier_total_balance(self.selected_supplier_id)
        
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"➕ {transaction_type} Ekle: {supplier['name']}")
        dialog.geometry("400x450")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        tk.Label(dialog, text=f"Toptancı: {supplier['name']}", font=("Arial", 11, "bold")).pack(pady=10)
        tk.Label(dialog, text=f"Mevcut Bakiye: ₺{current_balance:.2f}", font=("Arial", 10)).pack()
        
        # İşlem türü seçimi
        tk.Label(dialog, text="İşlem Türü:").pack(pady=5)
        type_combo = ttk.Combobox(dialog, values=["BORC", "ALACAK"], state="readonly", width=28)
        type_combo.set(transaction_type)
        type_combo.pack()
        
        tk.Label(dialog, text="Tutar:").pack(pady=5)
        amount_entry = tk.Entry(dialog, width=20, font=("Arial", 12))
        amount_entry.pack()
        
        tk.Label(dialog, text="Tarih:").pack(pady=5)
        date_entry = tk.Entry(dialog, width=20, font=("Arial", 11))
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        date_entry.pack()
        
        tk.Label(dialog, text="Vade Tarihi (isteğe bağlı):").pack(pady=5)
        due_date_entry = tk.Entry(dialog, width=20, font=("Arial", 11))
        due_date_entry.pack()
        
        tk.Label(dialog, text="Açıklama:").pack(pady=5)
        desc_text = tk.Text(dialog, height=3, width=40)
        desc_text.insert(tk.END, f"{transaction_type} eklendi")
        desc_text.pack()
        
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        tk.Button(
            button_frame,
            text="💾 Kaydet",
            command=lambda: self.save_balance_transaction(dialog, supplier['id'], type_combo, amount_entry, date_entry, due_date_entry, desc_text),
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
        
        tk.Button(
            button_frame,
            text="❌ İptal",
            command=dialog.destroy,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
    
    def save_balance_transaction(self, dialog, supplier_id, type_combo, amount_entry, date_entry, due_date_entry, desc_text):
        """Bakiye işlemini kaydeder"""
        try:
            transaction_type = type_combo.get()
            amount = float(amount_entry.get())
            transaction_date = date_entry.get().strip()
            due_date = due_date_entry.get().strip() or None
            description = desc_text.get("1.0", tk.END).strip()
            
            # Tarih doğrulama
            try:
                datetime.strptime(transaction_date, "%Y-%m-%d")
                if due_date:
                    datetime.strptime(due_date, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning("Hatalı Tarih", "Tarih formatı: YYYY-MM-DD")
                return
            
            if amount <= 0:
                messagebox.showwarning("Hatalı Değer", "Tutar 0'dan büyük olmalı!")
                return
            
            # Akıllı bakiye işlemi
            new_balance = add_smart_balance_transaction(
                supplier_id,
                transaction_type,
                amount,
                due_date,
                description,
                transaction_date
            )
            
            if new_balance is not None:
                supplier = fetch_one("SELECT * FROM suppliers WHERE id = ?", (supplier_id,))
                messagebox.showinfo("Başarılı", f"İşlem kaydedildi!\nYeni bakiye: ₺{new_balance:.2f}")
                dialog.destroy()
                self.load_balances()
                self.load_suppliers()
            else:
                messagebox.showerror("Hata", "İşlem kaydedilemedi!")
                
        except ValueError:
            messagebox.showwarning("Hatalı Değer", "Tutar sayısal olmalı!")
    
    def payment_from_balance(self):
        """Bakiye listesinden ödeme yap"""
        selection = self.balance_tree.selection()
        if not selection:
            messagebox.showwarning("Seçim Yok", "Lütfen bir bakiye seçin!")
            return
        
        values = self.balance_tree.item(selection[0])['values']
        
        # Sadece BORÇ bakiyelerde ödeme yapılabilir
        if values[2] != "BORC":
            messagebox.showinfo("Bilgi", "Sadece borç bakiyelerde ödeme yapılabilir!")
            return
        
        balance_id = fetch_one("""
            SELECT sb.id 
            FROM supplier_balances sb
            JOIN suppliers sup ON sb.supplier_id = sup.id
            WHERE sup.name = ? AND sb.date = ? AND sb.type = 'BORC'
            ORDER BY sb.id DESC LIMIT 1
        """, (values[1], values[0]))['id']
        
        if not balance_id:
            return
        
        self.payment_from_selected_balance(balance_id, values[1])
    
    def payment_from_selected_balance(self, balance_id, supplier_name):
        """Seçili bakiye için ödeme yap"""
        balance = fetch_one("SELECT * FROM supplier_balances WHERE id = ?", (balance_id,))
        current_balance = get_supplier_total_balance(balance['supplier_id'])
        
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"💵 Ödeme Yap: {supplier_name}")
        dialog.geometry("400x350")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        tk.Label(dialog, text=f"Toptancı: {supplier_name}", font=("Arial", 11, "bold")).pack(pady=10)
        tk.Label(dialog, text=f"Borç Tutarı: ₺{balance['amount']:.2f}", font=("Arial", 10)).pack()
        tk.Label(dialog, text=f"Mevcut Toplam Bakiye: ₺{current_balance:.2f}", font=("Arial", 10)).pack()
        
        tk.Label(dialog, text="Ödeme Tutarı:").pack(pady=10)
        amount_entry = tk.Entry(dialog, width=20, font=("Arial", 12))
        amount_entry.insert(0, str(balance['amount']))  # Varsayılan tam ödeme
        amount_entry.pack()
        
        tk.Label(dialog, text="Tarih:").pack(pady=5)
        date_entry = tk.Entry(dialog, width=20, font=("Arial", 11))
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        date_entry.pack()
        
        tk.Label(dialog, text="Açıklama:").pack(pady=5)
        desc_text = tk.Text(dialog, height=3, width=40)
        desc_text.insert(tk.END, "Ödeme yapıldı")
        desc_text.pack()
        
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        tk.Button(
            button_frame,
            text="💵 Ödeme Yap",
            command=lambda: self.process_smart_payment(dialog, balance['supplier_id'], amount_entry, date_entry, desc_text, current_balance),
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
        
        tk.Button(
            button_frame,
            text="❌ İptal",
            command=dialog.destroy,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
    
    def update_balance_status(self):
        """Bakiye durumunu manuel güncelle"""
        selection = self.balance_tree.selection()
        if not selection:
            messagebox.showwarning("Seçim Yok", "Lütfen bir bakiye seçin!")
            return
        
        values = self.balance_tree.item(selection[0])['values']
        balance_id = values[0]
        
        dialog = tk.Toplevel(self.parent)
        dialog.title("🔄 Durum Güncelle")
        dialog.geometry("300x150")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        tk.Label(dialog, text="Yeni Durum:").pack(pady=10)
        status_combo = ttk.Combobox(dialog, values=["AKTIF", "ODENDI", "GECIKMIS"], state="readonly")
        status_combo.set(values[6])  # Mevcut durumu göster
        status_combo.pack()
        
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        tk.Button(
            button_frame,
            text="💾 Güncelle",
            command=lambda: self.save_status(dialog, balance_id, status_combo),
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
        
        tk.Button(
            button_frame,
            text="❌ İptal",
            command=dialog.destroy,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=10)
    
    def save_status(self, dialog, balance_id, status_combo):
        """Durumu kaydeder"""
        from database import update_balance_status
        status = status_combo.get()
        
        success = update_balance_status(balance_id, status)
        
        if success:
            messagebox.showinfo("Başarılı", f"Bakiye durumu güncellendi: {status}")
            dialog.destroy()
            self.load_balances()
        else:
            messagebox.showerror("Hata", "Durum güncellenemedi!")
    
    def show_all_transactions(self):
        """Toptancının tüm hareketlerini göster"""
        if not self.selected_supplier_id:
            return
        
        supplier = fetch_one("SELECT * FROM suppliers WHERE id = ?", (self.selected_supplier_id,))
        transactions = get_supplier_transaction_history(self.selected_supplier_id)
        
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"📋 Tüm Hareketler: {supplier['name']}")
        dialog.geometry("800x500")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # Treeview
        frame = tk.Frame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        y_scroll = tk.Scrollbar(frame)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        tree = ttk.Treeview(
            frame,
            columns=("Date", "Type", "Amount", "Status", "Description"),
            show="headings",
            height=20,
            yscrollcommand=y_scroll.set
        )
        
        tree.heading("Date", text="Tarih")
        tree.heading("Type", text="İşlem Türü")
        tree.heading("Amount", text="Tutar")
        tree.heading("Status", text="Durum")
        tree.heading("Description", text="Açıklama")
        
        tree.column("Date", width=140)
        tree.column("Type", width=100, anchor="center")
        tree.column("Amount", width=100, anchor="e")
        tree.column("Status", width=80, anchor="center")
        tree.column("Description", width=350)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        y_scroll.config(command=tree.yview)
        
        # Verileri ekle
        for trans in transactions:
            tree.insert("", tk.END, values=(
                trans['date'],
                trans['type'],
                f"₺{trans['amount']:.2f}",
                trans['status'],
                trans['description'] or "-"
            ))
        
        tk.Button(dialog, text="✅ Tamam", command=dialog.destroy, bg="#4CAF50", fg="white").pack(pady=10)
    
    def show_balance_details(self):
        """Bakiye detaylarını göster"""
        selection = self.balance_tree.selection()
        if not selection:
            return
        
        values = self.balance_tree.item(selection[0])['values']
        
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"📋 Bakiye Detayı")
        dialog.geometry("500x300")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # Detayları göster
        text_widget = tk.Text(dialog, height=10, width=60, font=("Arial", 10))
        text_widget.pack(pady=10, padx=10)
        
        details = f"""
📋 BAKİYE DETAYI

Toptancı: {values[2]}
İşlem Türü: {values[3]}
Tutar: {values[4]}
Tarih: {values[1]}
Vade Tarihi: {values[5]}
Durum: {values[6]}
Açıklama: {values[7] or 'Yok'}

💡 İPUCU:
- Ödeme yapmak için sağ tık → Ödeme Yap
- Durum güncellemek için sağ tık → Durum Güncelle
"""
        
        text_widget.insert(tk.END, details)
        text_widget.config(state=tk.DISABLED)
        
        tk.Button(dialog, text="✅ Tamam", command=dialog.destroy, bg="#4CAF50", fg="white").pack(pady=10)
    
    def show_due_payments(self):
        """Yaklaşan ödemeler – vadesiz kayıtlara takılmaz"""
        from datetime import datetime   # 🔥 sade datetime

        try:
            due_balances = get_due_supplier_balances(self.branch_id, days=7)

            if not due_balances:
                messagebox.showinfo("Bilgi", "Yaklaşan ödeme bulunmuyor!")
                return

            dialog = tk.Toplevel(self.parent)
            dialog.title("⚠️ Yaklaşan Ödemeler (7 Gün)")
            dialog.geometry("800x500")
            dialog.transient(self.parent)
            dialog.grab_set()

            main_frame = tk.Frame(dialog)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            tk.Label(main_frame, text="⚠️ Yaklaşan Ödemeler (7 Gün İçinde)",
                     font=("Arial", 14, "bold"), fg="#d32f2f").pack(pady=10)

            tree_frame = tk.Frame(main_frame)
            tree_frame.pack(fill=tk.BOTH, expand=True, pady=10)

            y_scroll = tk.Scrollbar(tree_frame)
            y_scroll.pack(side=tk.RIGHT, fill=tk.Y)

            tree = ttk.Treeview(
                tree_frame,
                columns=("Supplier", "Type", "Amount", "DueDate", "DaysLeft"),
                show="headings",
                height=15,
                yscrollcommand=y_scroll.set
            )

            for col, text in zip(("Supplier", "Type", "Amount", "DueDate", "DaysLeft"),
                                 ("Toptancı", "Tür", "Tutar", "Vade Tarihi", "Kalan Gün")):
                tree.heading(col, text=text)
                tree.column(col, width=200 if col == "Supplier" else 100, anchor="center" if col != "Supplier" else "w")

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            y_scroll.config(command=tree.yview)

            total_amount = 0
            now = datetime.now()

            for balance in due_balances:
                # 🔥 NULL kontrollü tarih işlemi
                due_date = datetime.strptime(balance['due_date'], "%Y-%m-%d")
                delta_days = (due_date - now).days

                if delta_days < 0:
                    days_text, tag = f"🔴 Gecikti ({abs(delta_days)} gün)", 'late'
                elif delta_days == 0:
                    days_text, tag = "🔥 BUGÜN", 'warning'
                else:
                    days_text, tag = f"{delta_days} gün", 'caution' if delta_days <= 3 else 'active'

                tree.insert("", tk.END, values=(
                    balance['supplier_name'],
                    balance['type'],
                    f"₺{balance['amount']:.2f}",
                    balance['due_date'],
                    days_text
                ), tags=(tag,))

                total_amount += balance['amount']

            # Renk ayarları
            for t, bg, fg in [('late', '#ffebee', '#c62828'),
                              ('warning', '#fff3e0', '#e65100'),
                              ('caution', '#fff9c4', '#f57f17'),
                              ('active', '#e8f5e9', '#2e7d32')]:
                tree.tag_configure(t, background=bg, foreground=fg)

            tk.Label(main_frame, text=f"💰 Toplam Tutar: ₺{total_amount:.2f}",
                     font=("Arial", 12, "bold")).pack(pady=10)
            tk.Button(main_frame, text="✅ Tamam", command=dialog.destroy,
                     bg="#4CAF50", fg="white", font=("Arial", 10, "bold")).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Hata", f"Yaklaşan ödemeler getirilirken hata oluştu:\n{str(e)}")
    def export_to_excel(self):
        """Excel raporu oluştur - HATASIZ VERSİYON"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                "Eksik Kütüphane",
                "Excel aktarma için 'openpyxl' kütüphanesi gereklidir.\n\n"
                "pip install openpyxl"
            )
            return

        # Verileri getir
        suppliers = get_all_suppliers(self.branch_id)
        balances = get_supplier_balances(self.branch_id)

        if not suppliers:
            messagebox.showinfo("Bilgi", "Aktarılacak toptancı bulunmuyor!")
            return

        # Excel oluştur
        wb = Workbook()

        # Toptancı sayfası
        ws_suppliers = wb.active
        ws_suppliers.title = "Toptancı Listesi"

        # Stil tanımları
        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True)
        normal_font = Font(name="Calibri", size=10)

        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Başlık
        ws_suppliers['A1'] = "TOPTANCI RAPORU"
        ws_suppliers['A1'].font = title_font
        ws_suppliers.merge_cells('A1:E1')

        # Bilgiler
        ws_suppliers['A3'] = f"Rapor Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_suppliers['A4'] = f"Toplam Toptancı: {len(suppliers)}"

        # Toptancı tablosu başlıkları
        supplier_headers = ["ID", "Adı", "Türü", "Telefon", "Toplam Bakiye"]

        for col_idx, header in enumerate(supplier_headers, 1):
            cell = ws_suppliers.cell(row=6, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Toptancı verileri
        row_idx = 7
        for supplier in suppliers:
            total_balance = get_supplier_total_balance(supplier['id'])

            row_data = [
                supplier['id'],
                supplier['name'],
                supplier['supplier_type'],
                supplier['phone'] or "-",
                f"₺{total_balance:.2f}"
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws_suppliers.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

                # Renklendirme
                if col_idx == 5:
                    if total_balance > 0:
                        cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                    elif total_balance < 0:
                        cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")

            row_idx += 1

        # Bakiye detay sayfası
        ws_balances = wb.create_sheet("Bakiye Detayları")

        # Bakiye tablosu başlıkları
        balance_headers = ["Tarih", "Toptancı", "İşlem Türü", "Tutar", "Vade Tarihi", "Durum", "Açıklama"]

        for col_idx, header in enumerate(balance_headers, 1):
            cell = ws_balances.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Bakiye verileri
        row_idx = 2
        for balance in balances:
            row_data = [
                balance['date'],
                balance['supplier_name'],
                balance['balance_type'],
                f"₺{balance['amount']:.2f}",
                balance['due_date'] or "-",
                balance['status'],
                balance['description'] or "-"
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws_balances.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

                # Duruma göre renklendirme
                if col_idx == 6:
                    status = balance['status']
                    if status == 'GECIKMIS':
                        cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                    elif status == 'ODENDI':
                        cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                    elif status == 'AKTIF':
                        cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")

            row_idx += 1

        # 🔥 HATANIN OLDUĞU KISIM - DÜZELTİLDİ
        # Manuel genişlik ayarı (birleştirilmiş hücreler için güvenli)
        width_settings = {
            'A': 12, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 12, 'G': 25
        }
        
        for sheet in [ws_suppliers, ws_balances]:
            for col_letter, width in width_settings.items():
                if width_settings.get(col_letter):
                    sheet.column_dimensions[col_letter].width = width

        # Dosya kaydet
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"toptanci_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}"
        )

        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo(
                    "✅ Başarılı", 
                    f"Excel dosyası kaydedildi:\n{file_path}\n\n"
                    f"📊 {len(suppliers)} toptancı ve {len(balances)} bakiye kaydı aktarıldı."
                )
            except Exception as e:
                messagebox.showerror(
                    "❌ Hata",
                    f"Excel dosyası kaydedilemedi:\n{str(e)}"
                )
