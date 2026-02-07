# modules/finance_tab.py
import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime, timedelta
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import (
    add_transaction, get_transactions, get_daily_total, get_period_summary,
    update_transaction, delete_transaction, get_transaction_stats, fetch_one
)
from modules.ui_helpers import show_info, show_warning, show_error, ask_confirm, show_toast

class FinanceTab:
    def __init__(self, parent, branch_id):
        self.parent = parent
        self.branch_id = branch_id
        
        self.create_widgets()
        self.load_transactions()
        self.update_summary()
    
    def create_widgets(self):
        # Üst çerçeve - Kontroller
        top_frame = ttk.Frame(self.parent)
        top_frame.pack(fill=tk.X, padx=16, pady=12)
        
        # Sol taraf - Hızlı Ekleme
        left_frame = ttk.Frame(top_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 20))
        
        ttk.Label(left_frame, text="💰 Hızlı İşlem Ekle").pack(pady=5)
        
        # İşlem türü
        ttk.Label(left_frame, text="Tür:").pack()
        self.trans_type = tk.StringVar(value="GELIR")
        ttk.Radiobutton(left_frame, text="Gelir", variable=self.trans_type, value="GELIR").pack()
        ttk.Radiobutton(left_frame, text="Gider", variable=self.trans_type, value="GIDER").pack()
        
        # Miktar
        ttk.Label(left_frame, text="Miktar:").pack(pady=(10, 0))
        self.amount_entry = ttk.Entry(left_frame, width=15)
        self.amount_entry.pack()
        
        # Ödeme yöntemi
        ttk.Label(left_frame, text="Ödeme:").pack(pady=(10, 0))
        self.payment_var = tk.StringVar(value="NAKIT")
        payment_combo = ttk.Combobox(left_frame, textvariable=self.payment_var, 
                                    values=["NAKIT", "KREDI", "BANKA", "DIGER"], 
                                    state="readonly", width=13)
        payment_combo.pack()
        
        # Tarih
        ttk.Label(left_frame, text="Tarih:").pack(pady=(10, 0))
        self.date_entry = ttk.Entry(left_frame, width=15)
        self.date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.date_entry.pack()
        
        # Açıklama
        ttk.Label(left_frame, text="Açıklama:").pack(pady=(10, 0))
        self.desc_entry = tk.Text(left_frame, height=3, width=20, relief="solid", borderwidth=1)
        self.desc_entry.pack()
        
        # Ekle butonu
        ttk.Button(left_frame, text="➕ Ekle", command=self.add_transaction).pack(pady=10)
        
        # Sağ taraf - Kontroller
        right_frame = ttk.Frame(top_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Rapor butonları
        ttk.Label(right_frame, text="📊 Raporlar").pack(pady=5)
        
        ttk.Button(right_frame, text="Günlük Rapor", command=self.show_daily_report).pack(pady=2, fill=tk.X)
        ttk.Button(right_frame, text="Haftalık Rapor", command=self.show_weekly_report).pack(pady=2, fill=tk.X)
        ttk.Button(right_frame, text="Aylık Rapor", command=self.show_monthly_report).pack(pady=2, fill=tk.X)
        ttk.Button(right_frame, text="Özel Aralık", command=self.show_custom_range).pack(pady=2, fill=tk.X)
        ttk.Button(right_frame, text="📈 Excel'e Aktar", command=self.export_to_excel).pack(pady=2, fill=tk.X)
        
        # Orta kısım - Özet
        middle_frame = ttk.Frame(top_frame)
        middle_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20)
        
        self.summary_frame = ttk.LabelFrame(middle_frame, text="📈 Güncel Özet", padding=10)
        self.summary_frame.pack(fill=tk.BOTH, expand=True)
        
        self.summary_labels = {}
        for key, text in [('income', 'Günlük Gelir'), ('expense', 'Günlük Gider'), ('net', 'Net Kâr')]:
            frame = ttk.Frame(self.summary_frame)
            frame.pack(fill=tk.X, pady=5)
            ttk.Label(frame, text=f"{text}:").pack(side=tk.LEFT)
            self.summary_labels[key] = ttk.Label(frame, text="₺0.00")
            self.summary_labels[key].pack(side=tk.RIGHT)
        
        # Orta çerçeve - İşlem listesi
        list_frame = ttk.Frame(self.parent)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)
        
        # Üst kontroller
        control_frame = ttk.Frame(list_frame)
        control_frame.pack(fill=tk.X, pady=5)
        
        # Arama ve filtreleme
        ttk.Label(control_frame, text="🔍 Ara:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search_change)
        ttk.Entry(control_frame, textvariable=self.search_var, width=20).pack(side=tk.LEFT, padx=5)
        
        # Filtreler
        self.filter_type = tk.StringVar(value="Tümü")
        type_combo = ttk.Combobox(control_frame, textvariable=self.filter_type, 
                                 values=["Tümü", "GELIR", "GIDER"], state="readonly", width=10)
        type_combo.pack(side=tk.LEFT, padx=5)
        type_combo.bind("<<ComboboxSelected>>", lambda e: self.apply_filters())
        
        self.filter_payment = tk.StringVar(value="Tümü")
        payment_combo = ttk.Combobox(control_frame, textvariable=self.filter_payment, 
                                    values=["Tümü", "NAKIT", "KREDI", "BANKA", "DIGER"], 
                                    state="readonly", width=10)
        payment_combo.pack(side=tk.LEFT, padx=5)
        payment_combo.bind("<<ComboboxSelected>>", lambda e: self.apply_filters())
        
        # Tarih aralığı
        ttk.Label(control_frame, text="Tarih:").pack(side=tk.LEFT, padx=(20, 5))
        self.start_date = ttk.Entry(control_frame, width=10)
        self.start_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.start_date.pack(side=tk.LEFT)
        
        ttk.Label(control_frame, text="-").pack(side=tk.LEFT)
        self.end_date = ttk.Entry(control_frame, width=10)
        self.end_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.end_date.pack(side=tk.LEFT)
        
        ttk.Button(control_frame, text="🔍 Uygula", command=self.apply_filters).pack(side=tk.LEFT, padx=5)
        
        # Treeview
        self.tree = ttk.Treeview(list_frame, columns=("ID", "Date", "Type", "Amount", "Payment", "Description"), 
                                show="headings", height=15)
        
        # Başlıklar
        headers = [("ID", "ID", 50), ("Date", "Tarih", 120), ("Type", "Tür", 80), 
                  ("Amount", "Tutar", 100), ("Payment", "Ödeme", 80), ("Description", "Açıklama", 250)]
        
        for col, text, width in headers:
            self.tree.heading(col, text=text)
            self.tree.column(col, width=width, anchor="center" if col != "Description" else "w")
        
        self.tree.pack(fill=tk.BOTH, expand=True)

        self.empty_label = ttk.Label(list_frame, text="Henüz işlem yok. İlk işlemi ekleyin.", style="Empty.TLabel")
        self.empty_label.place(relx=0.5, rely=0.5, anchor="center")
        
        # Sağ tık menüsü
        self.context_menu = tk.Menu(self.tree, tearoff=0)
        self.context_menu.add_command(label="✏️ Düzenle", command=self.edit_transaction)
        self.context_menu.add_command(label="🗑️ Sil", command=self.delete_transaction)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="📋 Detaylar", command=self.show_details)
        
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<Double-1>", lambda e: self.edit_transaction())
        
        # Alt toplam
        bottom_frame = ttk.Frame(self.parent)
        bottom_frame.pack(fill=tk.X, padx=16, pady=12)
        
        self.total_label = ttk.Label(bottom_frame, text="Toplam: ₺0.00")
        self.total_label.pack(side=tk.LEFT)
        
        ttk.Button(bottom_frame, text="🔄 Yenile", command=self.refresh_all).pack(side=tk.RIGHT)
    
    def add_transaction(self):
        """Yeni işlem ekle"""
        try:
            amount = float(self.amount_entry.get())
        except ValueError:
            show_warning(self.parent, "Hatalı Değer", "Miktar sayısal olmalı!")
            return
        
        trans_type = self.trans_type.get()
        payment = self.payment_var.get()
        date = self.date_entry.get()
        description = self.desc_entry.get("1.0", tk.END).strip()
        
        # Tarih doğrulama
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            show_warning(self.parent, "Hatalı Tarih", "Tarih formatı: YYYY-MM-DD")
            return
        
        trans_id = add_transaction(self.branch_id, trans_type, amount, payment, date, description)
        
        if trans_id:
            show_toast(self.parent, "İşlem kaydedildi!")
            self.clear_form()
            self.load_transactions()
            self.update_summary()
        else:
            show_error(self.parent, "Hata", "İşlem kaydedilemedi!")
    
    def clear_form(self):
        """Formu temizle"""
        self.amount_entry.delete(0, tk.END)
        self.desc_entry.delete("1.0", tk.END)
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
    
    def load_transactions(self, filters=None):
        """İşlemleri yükle"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        if filters is None:
            filters = {}
        
        transactions = get_transactions(self.branch_id, **filters)
        
        # Arama filtresi
        search_term = self.search_var.get().lower()
        
        total = 0
        for trans in transactions:
            # Arama filtresi
            description = trans['description'] or ""
            if search_term and search_term not in description.lower() and search_term not in str(trans['amount']):
                continue
            
            self.tree.insert("", tk.END, values=(
                trans['id'],
                trans['date'][:10],  # Sadece tarih kısmı
                trans['type'],
                f"₺{trans['amount']:.2f}",
                trans['payment_method'],
                trans['description'] or "-"
            ))
            
            # Toplam hesaplama
            if trans['type'] == 'GELIR':
                total += trans['amount']
            else:
                total -= trans['amount']
        
        self.total_label.config(text=f"Toplam: ₺{total:.2f}")

        if len(self.tree.get_children()) == 0:
            self.empty_label.place(relx=0.5, rely=0.5, anchor="center")
        else:
            self.empty_label.place_forget()
    
    def apply_filters(self):
        """Filtreleri uygula"""
        filters = {}
        
        if self.filter_type.get() != "Tümü":
            filters['trans_type'] = self.filter_type.get()
        
        if self.filter_payment.get() != "Tümü":
            filters['payment_method'] = self.filter_payment.get()
        
        if self.start_date.get() and self.end_date.get():
            filters['start_date'] = self.start_date.get()
            filters['end_date'] = self.end_date.get()
        
        self.load_transactions(filters)
    
    def on_search_change(self, *args):
        """Arama değiştiğinde"""
        self.apply_filters()
    
    def show_context_menu(self, event):
        """Sağ tık menüsü"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)
    
    def edit_transaction(self):
        """İşlem düzenle"""
        selection = self.tree.selection()
        if not selection:
            show_warning(self.parent, "Seçim Yok", "Lütfen bir işlem seçin!")
            return
        
        values = self.tree.item(selection[0])['values']
        trans_id = values[0]
        
        # Transaction'ı getir
        trans = fetch_one("SELECT * FROM transactions WHERE id = ?", (trans_id,))
        if not trans:
            return
        
        dialog = tk.Toplevel(self.parent)
        dialog.title("✏️ İşlem Düzenle")
        dialog.geometry("400x500")
        dialog.configure(bg="#f5f7fb")
        dialog.transient(self.parent)
        dialog.grab_set()
        dialog.bind("<Escape>", lambda e: dialog.destroy())
        
        # Form
        ttk.Label(dialog, text="İşlem Türü:").pack(pady=5)
        trans_type = tk.StringVar(value=trans['type'])
        ttk.Radiobutton(dialog, text="Gelir", variable=trans_type, value="GELIR").pack()
        ttk.Radiobutton(dialog, text="Gider", variable=trans_type, value="GIDER").pack()
        
        ttk.Label(dialog, text="Miktar:").pack(pady=5)
        amount_entry = ttk.Entry(dialog, width=20)
        amount_entry.insert(0, trans['amount'])
        amount_entry.pack()
        
        ttk.Label(dialog, text="Ödeme Yöntemi:").pack(pady=5)
        payment_var = tk.StringVar(value=trans['payment_method'])
        payment_combo = ttk.Combobox(dialog, textvariable=payment_var, 
                                    values=["NAKIT", "KREDI", "BANKA", "DIGER"], 
                                    state="readonly", width=18)
        payment_combo.pack()
        
        ttk.Label(dialog, text="Tarih:").pack(pady=5)
        date_entry = ttk.Entry(dialog, width=20)
        date_entry.insert(0, trans['date'][:10])
        date_entry.pack()
        
        ttk.Label(dialog, text="Açıklama:").pack(pady=5)
        desc_entry = tk.Text(dialog, height=5, width=40, relief="solid", borderwidth=1)
        desc_entry.insert(tk.END, trans['description'] or "")
        desc_entry.pack()
        
        # Butonlar
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="💾 Kaydet", 
                 command=lambda: self.save_edited_transaction(dialog, trans_id, trans_type, amount_entry, 
                                                          payment_var, date_entry, desc_entry)
                 ).pack(side=tk.LEFT, padx=10)
        
        ttk.Button(button_frame, text="❌ İptal", command=dialog.destroy
                 ).pack(side=tk.LEFT, padx=10)

        dialog.bind(
            "<Return>",
            lambda e: self.save_edited_transaction(
                dialog, trans_id, trans_type, amount_entry, payment_var, date_entry, desc_entry
            ),
        )
    
    def save_edited_transaction(self, dialog, trans_id, trans_type, amount_entry, 
                               payment_var, date_entry, desc_entry):
        """Düzenlenmiş işlemi kaydet"""
        try:
            amount = float(amount_entry.get())
        except ValueError:
            show_warning(dialog, "Hatalı Değer", "Miktar sayısal olmalı!")
            return
        
        trans_type_val = trans_type.get()
        payment = payment_var.get()
        date = date_entry.get()
        description = desc_entry.get("1.0", tk.END).strip()
        
        # Tarih doğrulama
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            show_warning(dialog, "Hatalı Tarih", "Tarih formatı: YYYY-MM-DD")
            return
        
        success = update_transaction(trans_id, trans_type_val, amount, payment, date, description)
        
        if success:
            show_toast(self.parent, "İşlem güncellendi!")
            dialog.destroy()
            self.load_transactions()
            self.update_summary()
        else:
            show_error(dialog, "Hata", "İşlem güncellenemedi!")
    
    def delete_transaction(self):
        """İşlem sil"""
        selection = self.tree.selection()
        if not selection:
            show_warning(self.parent, "Seçim Yok", "Lütfen bir işlem seçin!")
            return
        
        values = self.tree.item(selection[0])['values']
        trans_id = values[0]
        
        if ask_confirm(self.parent, "Silme Onayı", "Bu işlemi silmek istediğinize emin misiniz?"):
            success = delete_transaction(trans_id)
            if success:
                show_toast(self.parent, "İşlem silindi!")
                self.load_transactions()
                self.update_summary()
            else:
                show_error(self.parent, "Hata", "İşlem silinemedi!")
    
    def show_details(self):
        """İşlem detayları"""
        selection = self.tree.selection()
        if not selection:
            return
        
        values = self.tree.item(selection[0])['values']
        trans_id = values[0]
        trans = fetch_one("SELECT * FROM transactions WHERE id = ?", (trans_id,))
        
        if not trans:
            return
        
        dialog = tk.Toplevel(self.parent)
        dialog.title("📋 İşlem Detayı")
        dialog.geometry("400x300")
        dialog.configure(bg="#f5f7fb")
        dialog.transient(self.parent)
        dialog.grab_set()
        dialog.bind("<Escape>", lambda e: dialog.destroy())
        
        text = tk.Text(dialog, height=10, width=50, font=("Segoe UI", 10), relief="solid", borderwidth=1)
        text.pack(pady=10, padx=10)
        
        details = f"""
İşlem Detayı
{'='*30}

ID: {trans['id']}
Tarih: {trans['date']}
Tür: {trans['type']}
Miktar: ₺{trans['amount']:.2f}
Ödeme: {trans['payment_method']}
Açıklama: {trans['description'] or 'Yok'}
Kategori: {trans['category'] or 'Yok'}
        """
        
        text.insert(tk.END, details)
        text.config(state=tk.DISABLED)
        
        ttk.Button(dialog, text="✅ Tamam", command=dialog.destroy).pack(pady=10)
    
    def update_summary(self):
        """Günlük özet güncelle"""
        daily = get_daily_total(self.branch_id)
        
        colors = {'income': '#2e7d32', 'expense': '#c62828', 'net': '#1565c0'}
        
        for key, value in daily.items():
            self.summary_labels[key].config(
                text=f"₺{value:.2f}",
                fg=colors[key]
            )
    
    def show_daily_report(self):
        """Günlük rapor"""
        today = datetime.now().strftime("%Y-%m-%d")
        self.show_period_report("Günlük Rapor", today, today)
    
    def show_weekly_report(self):
        """Haftalık rapor"""
        today = datetime.now()
        week_start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
        week_end = today.strftime("%Y-%m-%d")
        self.show_period_report("Haftalık Rapor", week_start, week_end)
    
    def show_monthly_report(self):
        """Aylık rapor"""
        today = datetime.now()
        month_start = today.replace(day=1).strftime("%Y-%m-%d")
        month_end = today.strftime("%Y-%m-%d")
        self.show_period_report("Aylık Rapor", month_start, month_end)
    
    def show_custom_range(self):
        """Özel tarih aralığı raporu"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("📅 Özel Tarih Aralığı")
        dialog.geometry("300x200")
        dialog.configure(bg="#f5f7fb")
        dialog.transient(self.parent)
        dialog.grab_set()
        dialog.bind("<Escape>", lambda e: dialog.destroy())
        
        ttk.Label(dialog, text="Başlangıç Tarihi:").pack(pady=5)
        start_entry = ttk.Entry(dialog, width=12)
        start_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        start_entry.pack()
        
        ttk.Label(dialog, text="Bitiş Tarihi:").pack(pady=5)
        end_entry = ttk.Entry(dialog, width=12)
        end_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        end_entry.pack()
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="📊 Rapor Göster",
                 command=lambda: [self.show_period_report("Özel Aralık Raporu", 
                                                         start_entry.get(), end_entry.get()),
                               dialog.destroy()]
                 ).pack()

        dialog.bind(
            "<Return>",
            lambda e: [self.show_period_report("Özel Aralık Raporu", start_entry.get(), end_entry.get()), dialog.destroy()],
        )
    
    def show_period_report(self, title, start_date, end_date):
        """Tarih aralığı raporu"""
        try:
            # Tarih doğrulama
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            show_warning(self.parent, "Hatalı Tarih", "Tarih formatı: YYYY-MM-DD")
            return
        
        summary = get_period_summary(self.branch_id, start_date, end_date)
        transactions = get_transactions(self.branch_id, start_date, end_date)
        
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"📊 {title}")
        dialog.geometry("900x600")
        dialog.configure(bg="#f5f7fb")
        dialog.transient(self.parent)
        dialog.grab_set()
        dialog.bind("<Escape>", lambda e: dialog.destroy())
        
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Üst özet
        summary_frame = ttk.LabelFrame(main_frame, text=f"📈 {title} Özeti ({start_date} - {end_date})")
        summary_frame.pack(fill=tk.X, pady=10)
        
        for key, text, color in [('income', 'Toplam Gelir', '#2e7d32'),
                                ('expense', 'Toplam Gider', '#c62828'),
                                ('net', 'Net Kâr/Zarar', '#1565c0'),
                                ('count', 'İşlem Sayısı', '#607D8B')]:
            val = summary[key]
            display = f"₺{val:.2f}" if key != 'count' else str(int(val))
            ttk.Label(summary_frame, 
                    text=f"{text}: {display}",
                    foreground=color).pack(pady=5)
        
        # Liste
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        y_scroll = tk.Scrollbar(list_frame)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        tree = ttk.Treeview(list_frame, columns=("Date", "Type", "Amount", "Payment", "Description"),
                           show="headings", height=20, yscrollcommand=y_scroll.set)
        
        headers = [("Date", "Tarih", 120), ("Type", "Tür", 80), ("Amount", "Tutar", 100),
                  ("Payment", "Ödeme", 80), ("Description", "Açıklama", 400)]
        
        for col, text, width in headers:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center" if col != "Description" else "w")
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        y_scroll.config(command=tree.yview)
        
        # Verileri ekle
        total = 0
        for trans in transactions:
            tree.insert("", tk.END, values=(
                trans['date'][:10],
                trans['type'],
                f"₺{trans['amount']:.2f}",
                trans['payment_method'],
                trans['description'] or "-"
            ))
            total += trans['amount'] if trans['type'] == 'GELIR' else -trans['amount']
        
        # Alt toplam
        ttk.Label(main_frame, text=f"💰 Net Toplam: ₺{total:.2f}",
                foreground="#1565c0").pack(pady=10)
        
        ttk.Button(main_frame, text="✅ Tamam", command=dialog.destroy).pack(pady=10)
    
    def export_to_excel(self):
        """Excel raporu oluştur – HATASIZ"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            show_error(self.parent, "Eksik Kütüphane", "pip install openpyxl")
            return

        # Mevcut filtreleri al
        filters = {}
        if self.filter_type.get() != "Tümü":
            filters['trans_type'] = self.filter_type.get()
        if self.filter_payment.get() != "Tümü":
            filters['payment_method'] = self.filter_payment.get()
        if self.start_date.get() and self.end_date.get():
            filters['start_date'] = self.start_date.get()
            filters['end_date'] = self.end_date.get()

        transactions = get_transactions(self.branch_id, **filters)

        if not transactions:
            show_info(self.parent, "Bilgi", "Aktarılacak işlem bulunmuyor!")
            return

        wb = Workbook()

        # 1. Sayfa: Liste
        ws = wb.active
        ws.title = "Gelir_Gider_Raporu"  # 🔥 BURASI DEĞİŞTİ (/) yerine (_)

        # Stil
        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True)
        normal_font = Font(name="Calibri", size=10)

        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

        # Başlık
        ws['A1'] = "GELİR / GİDER RAPORU"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')

        # Tarih aralığı bilgisi
        date_range = f"{self.start_date.get()} - {self.end_date.get()}" if filters else "Tüm Zamanlar"
        ws['A3'] = f"Tarih Aralığı: {date_range}"
        ws['A4'] = f"Rapor Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Tablo başlıkları
        headers = ["ID", "Tarih", "Tür", "Tutar", "Ödeme", "Açıklama"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Veriler
        row_idx = 7
        for trans in transactions:
            row_data = [
                trans['id'],
                trans['date'][:10],
                trans['type'],
                f"₺{trans['amount']:.2f}",
                trans['payment_method'],
                trans['description'] or "-"
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

                # Renklendirme
                if col_idx == 4:  # Tutar
                    if trans['type'] == 'GELIR':
                        cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")

            row_idx += 1

        # 2. Sayfa: Özet
        ws_summary = wb.create_sheet("Özet")
        ws_summary.append(["Tarih Aralığı", date_range])
        ws_summary.append(["Toplam Gelir", f"₺{sum(t['amount'] for t in transactions if t['type'] == 'GELIR'):.2f}"])
        ws_summary.append(["Toplam Gider", f"₺{sum(t['amount'] for t in transactions if t['type'] == 'GIDER'):.2f}"])
        ws_summary.append(["Net Kâr/Zarar", f"₺{sum(t['amount'] if t['type'] == 'GELIR' else -t['amount'] for t in transactions):.2f}"])
        ws_summary.append(["İşlem Sayısı", len(transactions)])

        # Otomatik genişlik
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        for col in range(1, 3):
            ws_summary.column_dimensions[chr(64 + col)].width = 25

        # Dosya kaydet
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"gelir_gider_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}"
        )

        if file_path:
            wb.save(file_path)
            show_info(self.parent, "✅ Başarılı", f"Excel dosyası kaydedildi:\n{file_path}")
