# modules/stock_reports.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import get_all_products, get_stock_movements_report

class StockReportsDialog:
    def __init__(self, parent, branch_id):
        self.parent = parent
        self.branch_id = branch_id
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("📊 Stok Hareket Raporları")
        self.dialog.geometry("1000x650")
        self.dialog.configure(bg="#f5f7fb")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self.create_widgets()
        self.load_report()
    
    def create_widgets(self):
        # Üst çerçeve - Filtreler
        filter_frame = ttk.LabelFrame(self.dialog, text="Filtreler", padding=10)
        filter_frame.pack(fill=tk.X, padx=16, pady=12)
        
        # Tarih aralığı
        ttk.Label(filter_frame, text="Başlangıç:").grid(row=0, column=0, padx=5, pady=5)
        self.start_date = ttk.Entry(filter_frame, width=12)
        self.start_date.insert(0, (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        self.start_date.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(filter_frame, text="Bitiş:").grid(row=0, column=2, padx=5, pady=5)
        self.end_date = ttk.Entry(filter_frame, width=12)
        self.end_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.end_date.grid(row=0, column=3, padx=5, pady=5)
        
        # Ürün seçimi
        ttk.Label(filter_frame, text="Ürün:").grid(row=0, column=4, padx=5, pady=5)
        self.product_combo = ttk.Combobox(filter_frame, width=25, state="readonly")
        self.product_combo.grid(row=0, column=5, padx=5, pady=5)
        
        self.load_products_combo()
        
        # Butonlar
        ttk.Button(
            filter_frame,
            text="🔍 Uygula",
            command=self.load_report
        ).grid(row=0, column=6, padx=10)
        
        ttk.Button(
            filter_frame,
            text="📊 Excel'e Aktar",
            command=self.export_to_excel
        ).grid(row=0, column=7, padx=5)
        
        # İstatistik çerçevesi
        stats_frame = ttk.Frame(filter_frame)
        stats_frame.grid(row=1, column=0, columnspan=8, sticky="w", pady=(10, 0))
        
        self.total_movements_label = ttk.Label(stats_frame, text="Toplam Hareket: 0")
        self.total_movements_label.pack(side=tk.LEFT, padx=10)
        
        self.total_in_label = ttk.Label(stats_frame, text="Giriş: 0")
        self.total_in_label.pack(side=tk.LEFT, padx=10)
        
        self.total_out_label = ttk.Label(stats_frame, text="Çıkış: 0")
        self.total_out_label.pack(side=tk.LEFT, padx=10)
        
        # Orta çerçeve - Rapor tablosu
        list_frame = ttk.Frame(self.dialog)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)
        
        # Scrollbar'lar
        y_scroll = tk.Scrollbar(list_frame)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        x_scroll = tk.Scrollbar(list_frame, orient=tk.HORIZONTAL)
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Treeview
        self.tree = ttk.Treeview(
            list_frame,
            columns=("Date", "Product", "Barcode", "Type", "Qty", "OldQty", "NewQty", "Note"),
            show="headings",
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )
        
        # Kolon başlıkları
        self.tree.heading("Date", text="Tarih")
        self.tree.heading("Product", text="Ürün")
        self.tree.heading("Barcode", text="Barkod")
        self.tree.heading("Type", text="İşlem")
        self.tree.heading("Qty", text="Miktar")
        self.tree.heading("OldQty", text="Eski Miktar")
        self.tree.heading("NewQty", text="Yeni Miktar")
        self.tree.heading("Note", text="Not")
        
        # Kolon genişlikleri
        self.tree.column("Date", width=140)
        self.tree.column("Product", width=200)
        self.tree.column("Barcode", width=120)
        self.tree.column("Type", width=80, anchor="center")
        self.tree.column("Qty", width=80, anchor="center")
        self.tree.column("OldQty", width=80, anchor="center")
        self.tree.column("NewQty", width=80, anchor="center")
        self.tree.column("Note", width=200)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        y_scroll.config(command=self.tree.yview)
        x_scroll.config(command=self.tree.xview)
        
        # Alt çerçeve - Özet
        summary_frame = ttk.LabelFrame(self.dialog, text="Özet", padding=10)
        summary_frame.pack(fill=tk.X, padx=16, pady=12)
        
        self.summary_text = tk.Text(summary_frame, height=5, font=("Segoe UI", 9), wrap=tk.WORD, relief="solid", borderwidth=1)
        self.summary_text.pack(fill=tk.X)
        
        # Not
        ttk.Label(
            summary_frame,
            text="💡 Not: Miktar değişiklikleri için stok ekle/çıkar kullanın.",
            foreground="#64748b"
        ).pack(pady=(5, 0))
    
    def load_products_combo(self):
        """Ürün combobox'ını doldurur"""
        products = get_all_products(self.branch_id)
        
        combo_values = [("Tüm Ürünler", 0)]
        for product in products:
            combo_values.append((f"{product['name']} ({product['barcode'] or 'No Barcode'})", product['id']))
        
        self.product_combo['values'] = [val[0] for val in combo_values]
        self.product_combo.current(0)
        
        # Değerleri sakla
        self.combo_values = dict(combo_values)
    
    def load_report(self):
        """Raporu yükler"""
        start_date = self.start_date.get()
        end_date = self.end_date.get()
        
        # Tarih doğrulama
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            messagebox.showwarning("Hatalı Tarih", "Tarih formatı: YYYY-MM-DD")
            return
        
        # Seçili ürün
        selected_text = self.product_combo.get()
        product_id = self.combo_values.get(selected_text, 0)
        
        # Raporu getir
        movements = get_stock_movements_report(
            self.branch_id,
            product_id if product_id > 0 else None,
            start_date,
            end_date
        )
        
        # Treeview'i temizle
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # İstatistik değişkenleri
        total_in = 0
        total_out = 0
        
        # Verileri ekle
        for movement in movements:
            move_type = movement['type']
            qty = movement['quantity']
            
            if move_type == "IN":
                total_in += qty
                type_text = "GİRİŞ"
                tag = 'in'
            else:
                total_out += qty
                type_text = "ÇIKIŞ"
                tag = 'out'
            
            self.tree.insert("", tk.END, values=(
                movement['date'],
                movement['product_name'],
                movement['barcode'] or "-",
                type_text,
                qty,
                movement['old_quantity'],
                movement['new_quantity'],
                movement['note'] or "-"
            ), tags=(tag,))
        
        # Renklendirme
        self.tree.tag_configure('in', foreground='#4CAF50')
        self.tree.tag_configure('out', foreground='#c62828')
        
        # İstatistikleri güncelle
        total_movements = len(movements)
        self.total_movements_label.config(text=f"Toplam Hareket: {total_movements}")
        self.total_in_label.config(text=f"Giriş: {total_in} adet")
        self.total_out_label.config(text=f"Çıkış: {total_out} adet")
        
        # Özet metni güncelle
        self.update_summary(movements, total_in, total_out, start_date, end_date, selected_text)
    
    def update_summary(self, movements, total_in, total_out, start_date, end_date, selected_product):
        """Özet metnini günceller"""
        self.summary_text.delete("1.0", tk.END)
        
        if not movements:
            self.summary_text.insert(tk.END, "Seçilen kriterlere uygun hareket bulunamadı.")
            return
        
        # Genel özet
        self.summary_text.insert(tk.END, f"📊 STOK HAREKET RAPORU ÖZETİ\n")
        self.summary_text.insert(tk.END, f"{'='*50}\n\n")
        
        # Tarih aralığı
        self.summary_text.insert(tk.END, f"📅 Tarih Aralığı: {start_date} - {end_date}\n\n")
        
        # Ürün bilgisi
        if selected_product != "Tüm Ürünler":
            self.summary_text.insert(tk.END, f"📦 Ürün: {selected_product}\n\n")
        
        # Hareket istatistikleri
        self.summary_text.insert(tk.END, f"📈 Toplam Giriş: {total_in} adet\n")
        self.summary_text.insert(tk.END, f"📉 Toplam Çıkış: {total_out} adet\n")
        self.summary_text.insert(tk.END, f"📋 Net Değişim: {total_in - total_out} adet\n\n")
        
        # En çok hareket gören ürünler
        product_counts = {}
        for movement in movements:
            prod_name = movement['product_name']
            product_counts[prod_name] = product_counts.get(prod_name, 0) + 1
        
        if product_counts:
            self.summary_text.insert(tk.END, f"🔥 En Çok Hareket Gören Ürünler:\n")
            top_products = sorted(product_counts.items(), key=lambda x: x[1], reverse=True)[:5]
            for i, (prod, count) in enumerate(top_products, 1):
                self.summary_text.insert(tk.END, f"   {i}. {prod}: {count} hareket\n")
    
    def export_to_excel(self):
        """🚀 Raporu Excel'e aktarır"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Eksik Kütüphane",
                "Excel aktarma için 'openpyxl' kütüphanesi gereklidir.\n\n"
                "pip install openpyxl"
            )
            return
        
        # Mevcut filtreleri al
        start_date = self.start_date.get()
        end_date = self.end_date.get()
        selected_text = self.product_combo.get()
        product_id = self.combo_values.get(selected_text, 0)
        
        # Verileri getir
        movements = get_stock_movements_report(
            self.branch_id,
            product_id if product_id > 0 else None,
            start_date,
            end_date
        )
        
        if not movements:
            messagebox.showinfo("Bilgi", "Aktarılacak veri bulunamadı!")
            return
        
        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Stok Hareket Raporu"
        
        # Stil tanımları
        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True)
        normal_font = Font(name="Calibri", size=10)
        
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık
        ws['A1'] = "STOK HAREKET RAPORU"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        
        # Filtre bilgileri
        ws['A3'] = f"Rapor Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Tarih Aralığı: {start_date} - {end_date}"
        ws['A5'] = f"Ürün: {selected_text}"
        
        # Tablo başlıkları
        headers = ["Tarih", "Ürün Adı", "Barkod", "İşlem", "Miktar", "Eski Miktar", "Yeni Miktar", "Not"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Verileri ekle
        row_idx = 8
        for movement in movements:
            move_type = movement['type']
            row_data = [
                movement['date'],
                movement['product_name'],
                movement['barcode'] or "-",
                "GİRİŞ" if move_type == "IN" else "ÇIKIŞ",
                movement['quantity'],
                movement['old_quantity'],
                movement['new_quantity'],
                movement['note'] or "-"
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = green_fill if move_type == "IN" else red_fill
            
            row_idx += 1
        
        # Otomatik genişlik ayarı
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=7, max_row=row_idx, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Dosya kaydetme
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"stok_hareket_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}"
        )
        
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo(
                    "✅ Başarılı", 
                    f"Excel dosyası kaydedildi:\n{file_path}\n\n"
                    f"📊 Toplam {len(movements)} kayıt aktarıldı."
                )
            except Exception as e:
                messagebox.showerror(
                    "❌ Hata",
                    f"Excel dosyası kaydedilemedi:\n{str(e)}"
                )
